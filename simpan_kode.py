# -*- coding: utf-8 -*-
"""
Simpan kode "Produk Top 10" yang sudah dibetulkan di Excel.

Alurnya:
  1. Buka Excel hasil, sheet RINGKASAN
  2. Betulkan kolom **Kode** (yang kuning). Kolom "Produk Top 10" di
     SIAP_PASTE ikut berubah sendiri karena isinya rumus.
  3. SIMPAN Excel-nya
  4. Jalankan skrip ini -- kodenya masuk ke kode_produk.json

Tanpa langkah 4, kode yang kamu betulkan hilang bulan depan: generatornya
akan menebak lagi dari nama produk, dan tebakan itu hampir selalu salah.

  python simpan_kode.py                 <- baca semua Excel di hasil/
  python simpan_kode.py <file.xlsx>     <- satu file saja
"""
import glob
import os
import sys

from openpyxl import load_workbook

import konfigurasi as K
import lokasi as L

KOL_KODE = 2        # kolom B di sheet RINGKASAN
KOL_PRODUCT_ID = 4  # kolom D
BARIS_DATA = 6      # data mulai baris 6


def baca_kode(berkas):
    """Ambil {product_id: kode} dari sheet RINGKASAN satu file Excel."""
    try:
        wb = load_workbook(berkas, data_only=True)
    except Exception as e:
        print(f"  [x] tidak bisa dibuka: {e}")
        return {}
    if "RINGKASAN" not in wb.sheetnames:
        print("  [x] tidak ada sheet RINGKASAN, dilewati")
        return {}

    wr = wb["RINGKASAN"]
    hasil = {}
    for r in range(BARIS_DATA, wr.max_row + 1):
        pid = wr.cell(r, KOL_PRODUCT_ID).value
        kode = wr.cell(r, KOL_KODE).value
        if not pid or not kode:
            continue
        pid, kode = str(pid).strip(), str(kode).strip()
        if pid.isdigit() and kode:
            hasil[pid] = kode
    return hasil


def main():
    berkas = [a for a in sys.argv[1:] if a.lower().endswith(".xlsx")]
    if not berkas:
        berkas = sorted(glob.glob(os.path.join(L.DIR_HASIL, "*.xlsx")))
    if not berkas:
        print("Tidak ada file Excel di folder hasil/. Buat Excel-nya dulu.")
        return 1

    lama = dict(K.muat_kode_produk())
    baru = dict(lama)
    berubah, tambah = [], []

    for f in berkas:
        print(f"=== {os.path.basename(f)}")
        for pid, kode in baca_kode(f).items():
            if pid not in lama:
                tambah.append((pid, kode))
            elif lama[pid] != kode:
                berubah.append((pid, lama[pid], kode))
            baru[pid] = kode

    for pid, kode in tambah:
        print(f"  + {pid}  ->  {kode}")
    for pid, dari, ke in berubah:
        print(f"  ~ {pid}  {dari}  ->  {ke}")
    if not tambah and not berubah:
        print("\nTidak ada yang berubah. kode_produk.json sudah sesuai Excel.")
        return 0

    K.simpan_kode_produk(baru)
    print(f"\n[OK] {len(tambah)} kode baru, {len(berubah)} diperbarui")
    print(f"[OK] Tersimpan: {L.BERKAS_KODE}")
    print("\nBulan depan kode ini dipakai otomatis, tidak perlu dibetulkan lagi.")
    print("Supaya orang kantor ikut dapat, commit + push kode_produk.json.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

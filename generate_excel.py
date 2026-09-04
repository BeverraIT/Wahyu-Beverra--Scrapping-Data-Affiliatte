# -*- coding: utf-8 -*-
"""
LANGKAH 3.

Ubah JSON mentah (hasil/mentah/*.json) menjadi satu file Excel per toko:
  * Sheet SIAP_PASTE  -> 10 produk x 50 creator = 500 baris, tinggal blok-copy
    ke tab toko + bulan di spreadsheet AFFILIATE 3
  * Sheet RINGKASAN   -> cek cepat: tiap produk dapat berapa creator
  * Sheet 01..10      -> detail per produk (50 creator masing-masing)

Jalankan: python generate_excel.py
"""
import glob
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import konfigurasi as K

BIRU = "1F3864"
BIRU_MUDA = "D9E2F3"
ABU = "F2F2F2"
GARIS = Side(style="thin", color="BFBFBF")
BINGKAI = Border(left=GARIS, right=GARIS, top=GARIS, bottom=GARIS)


def _telusuri(baris, path):
    """Jalur bertitik: 'creator_meta.alias_name'."""
    cur = baris
    for bagian in path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(bagian)
        if cur is None:
            return None
    return cur


def ambil_nilai(baris, kandidat):
    """Ambil jalur pertama yang ada isinya. Nilai uang datang sebagai objek
    {'amount': '110276190', 'amount_formatted': 'Rp110.276.190', ...} --
    yang dipakai 'amount' supaya di Excel jadi angka, bukan teks."""
    for k in kandidat:
        v = _telusuri(baris, k)
        if isinstance(v, dict):
            v = v.get("amount", v.get("amount_formatted"))
        if v not in (None, "", []):
            return v
    return ""


IZIN_ANGKA = set("0123456789.,-+ ")


def angka(nilai):
    """Ubah 'Rp1.234.567' / '1,234' jadi angka. Teks biasa dibiarkan apa adanya
    supaya nama/username tidak ikut dipereteli."""
    if isinstance(nilai, (int, float)):
        return nilai
    teks = str(nilai).strip()
    inti = teks.lstrip("Rp").lstrip("IDR").strip()
    if not inti or not any(c.isdigit() for c in inti):
        return nilai
    if not set(inti) <= IZIN_ANGKA:
        return nilai
    # 1.234.567 (pemisah ribuan ID) vs 1234.56 (desimal)
    bersih = inti.replace(" ", "")
    if bersih.count(".") > 1 or (bersih.count(".") == 1 and len(bersih.split(".")[-1]) == 3):
        bersih = bersih.replace(".", "")
    bersih = bersih.replace(",", ".")
    try:
        return float(bersih) if "." in bersih else int(bersih)
    except Exception:
        return nilai


KOLOM_TETAP = ["No", "Peringkat Produk", "Nama Produk", "Product ID", "Peringkat Creator"]

# Kolom yang WAJIB tetap teks. Creator ID / Product ID itu 19 digit -- lewat
# batas presisi angka Excel (2^53), kalau ditulis sebagai angka digit
# belakangnya berubah jadi nol dan ID-nya rusak diam-diam.
KOLOM_TEKS = {"Creator ID", "Product ID", "Username", "Nama Creator", "Nama Produk"}

KUNING = "FFE699"      # warna header AFFILIATE 3
FORMAT_RUPIAH = '"Rp"#,##0'
FORMAT_ANGKA = "#,##0"


def nilai_kolom(judul, baris, kandidat):
    v = ambil_nilai(baris, kandidat)
    if judul in KOLOM_TEKS:
        return "" if v in (None, "") else str(v)
    return angka(v)


def kode_produk(prod):
    """Kode pendek untuk kolom 'Produk Top 10'. Kalau belum diisi tangan di
    KODE_PRODUK, pakai kata terakhir nama produk -- itu tebakan, bukan pasti."""
    kode = (K.KODE_PRODUK.get(str(prod.get("product_id"))) or "").strip()
    if kode:
        return kode, True
    kata = str(prod.get("nama_produk") or "").split()
    return (kata[-1].upper() if kata else ""), False


# Baris 1-4 RINGKASAN dipakai judul/keterangan, baris 5 header, data mulai 6.
BARIS_KEPALA_RINGKASAN = 5


def rujukan_kode(peringkat):
    """Rumus yang menunjuk sel Kode di RINGKASAN.

    Kolom "Produk Top 10" di SIAP_PASTE sengaja BUKAN nilai tetap: kode
    produk sering perlu dibetulkan tangan (yang ditebak dari nama produk
    hampir selalu salah). Dengan rumus ini, cukup ubah satu sel di RINGKASAN
    dan 50 baris produk itu ikut berubah."""
    return f"=RINGKASAN!$B${BARIS_KEPALA_RINGKASAN + peringkat}"


def nilai_siap_paste(sumber, creator, nomor, kode, tipe):
    if sumber == "@no":
        return nomor
    if sumber == "@kode":
        return kode
    if sumber == "":
        return ""
    if sumber.startswith("="):
        return sumber[1:]
    if tipe == "rupiah" and K.RUPIAH_SEBAGAI_TEKS:
        # ambil "Rp267.603.178" jadi-jadian dari API, bukan angkanya
        jadi = _telusuri(creator, sumber.rsplit(".", 1)[0] + ".amount_formatted")
        if jadi:
            return jadi
    return ambil_nilai(creator, [sumber])


def tulis_siap_paste(ws, r, prod, creator, nomor, kode):
    for kol, (judul, sumber, tipe) in enumerate(K.KOLOM_SIAP_PASTE, 1):
        v = nilai_siap_paste(sumber, creator, nomor, kode, tipe)
        teks_rupiah = tipe == "rupiah" and K.RUPIAH_SEBAGAI_TEKS
        if tipe in ("angka", "rupiah") and not teks_rupiah:
            v = angka(v)
            if v in (None, ""):
                v = 0
        else:
            v = "" if v in (None, "") else str(v)
        sel = ws.cell(row=r, column=kol, value=v)
        sel.border = BINGKAI
        sel.font = Font(size=10)
        sel.alignment = Alignment(horizontal="center", vertical="center")
        if tipe == "rupiah" and not teks_rupiah:
            sel.number_format = FORMAT_RUPIAH
        elif tipe == "angka":
            sel.number_format = FORMAT_ANGKA


def gaya_header(ws, jumlah_kolom, baris=1, warna=BIRU, teks="FFFFFF"):
    for c in range(1, jumlah_kolom + 1):
        sel = ws.cell(row=baris, column=c)
        sel.font = Font(bold=True, color=teks, size=10)
        sel.fill = PatternFill("solid", fgColor=warna)
        sel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sel.border = BINGKAI
    ws.row_dimensions[baris].height = 28
    ws.freeze_panes = ws.cell(row=baris + 1, column=1)


def lebar_otomatis(ws, maks=42):
    for kol in ws.columns:
        panjang = max((len(str(s.value)) for s in kol if s.value is not None), default=8)
        ws.column_dimensions[get_column_letter(kol[0].column)].width = min(max(panjang + 2, 10), maks)


def tulis_baris(ws, r, nilai, warnai_selang=True):
    for i, v in enumerate(nilai, 1):
        sel = ws.cell(row=r, column=i, value=v)
        sel.border = BINGKAI
        sel.font = Font(size=10)
        if warnai_selang and r % 2 == 0:
            sel.fill = PatternFill("solid", fgColor=ABU)
        if isinstance(v, (int, float)):
            sel.number_format = "#,##0"
            sel.alignment = Alignment(horizontal="right")


def bangun(berkas_json):
    with open(berkas_json, encoding="utf-8") as f:
        data = json.load(f)

    kolom_data = list(K.PETA_KOLOM.keys())

    wb = Workbook()

    # ---------- SIAP_PASTE (susunan persis AFFILIATE 3) ----------
    ws = wb.active
    ws.title = "SIAP_PASTE"
    ws.append([j for j, _, _ in K.KOLOM_SIAP_PASTE])
    gaya_header(ws, len(K.KOLOM_SIAP_PASTE), warna=KUNING, teks="000000")

    no = 0
    ringkasan = []
    kode_tebakan = []
    for urut, prod in enumerate(data["produk"], 1):
        creator = prod["creator"][:K.JUMLAH_CREATOR_PER_PRODUK]
        kode, pasti = kode_produk(prod)
        if not pasti:
            kode_tebakan.append((prod["peringkat"], prod["product_id"], kode))
        ringkasan.append((prod["peringkat"], kode, prod["nama_produk"],
                          prod["product_id"], len(creator)))
        # urut (bukan peringkat) supaya selalu cocok dengan urutan baris
        # yang benar-benar ditulis ke RINGKASAN di bawah
        rujukan = rujukan_kode(urut)
        for j, c in enumerate(creator, 1):
            no += 1
            tulis_siap_paste(ws, no + 1, prod, c, j, rujukan)
    lebar_otomatis(ws)

    # ---------- RINGKASAN ----------
    wr = wb.create_sheet("RINGKASAN", 0)
    wr["A1"] = f"{data['toko']} — {data['periode']}"
    wr["A1"].font = Font(bold=True, size=14, color=BIRU)
    wr["A2"] = f"Periode: {data['awal']} s/d {data['akhir']}"
    wr["A3"] = f"Total baris siap paste: {no} (target {K.JUMLAH_PRODUK_TOP * K.JUMLAH_CREATOR_PER_PRODUK})"
    wr["A3"].font = Font(bold=True, color="C00000" if no < K.JUMLAH_PRODUK_TOP * K.JUMLAH_CREATOR_PER_PRODUK else "1E7145")
    kepala = ["Peringkat", "Kode", "Nama Produk", "Product ID", "Jumlah Creator", "Status"]
    wr.append([])
    wr.append(kepala)
    gaya_header(wr, len(kepala), baris=5)
    for i, (pk, kd, nm, pid, jml) in enumerate(ringkasan, BARIS_KEPALA_RINGKASAN + 1):
        status = "LENGKAP" if jml >= K.JUMLAH_CREATOR_PER_PRODUK else f"KURANG {K.JUMLAH_CREATOR_PER_PRODUK - jml}"
        tulis_baris(wr, i, [pk, kd, nm, pid, jml, status])
        wr.cell(row=i, column=4).number_format = "@"
        # Kolom Kode ditandai kuning: ini satu-satunya sel yang boleh diubah
        # tangan, dan SIAP_PASTE mengikutinya lewat rumus.
        sel_kode = wr.cell(row=i, column=2)
        sel_kode.fill = PatternFill("solid", fgColor=KUNING)
        sel_kode.font = Font(size=10, bold=True)
        sel_kode.alignment = Alignment(horizontal="center")
        if status != "LENGKAP":
            wr.cell(row=i, column=6).fill = PatternFill("solid", fgColor="FFC7CE")

    baris_catatan = BARIS_KEPALA_RINGKASAN + len(ringkasan) + 2
    wr.cell(row=baris_catatan, column=1,
            value="Kolom Kode (kuning) boleh diubah. Kolom \"Produk Top 10\" di "
                  "SIAP_PASTE ikut berubah sendiri.").font = Font(bold=True, color=BIRU)
    wr.cell(row=baris_catatan + 1, column=1,
            value="Waktu menyalin ke AFFILIATE 3, tempel sebagai NILAI "
                  "(Paste Special > Values / Ctrl+Shift+V) supaya rumusnya "
                  "tidak ikut terbawa.").font = Font(color="C00000")
    lebar_otomatis(wr)

    # ---------- per produk ----------
    for prod in data["produk"]:
        wp = wb.create_sheet(f"{prod['peringkat']:02d}")
        wp["A1"] = prod["nama_produk"]
        wp["A1"].font = Font(bold=True, size=12, color=BIRU)
        wp["A2"] = f"Product ID: {prod['product_id']}"
        wp.append([])
        kepala2 = ["Peringkat Creator"] + kolom_data
        wp.append(kepala2)
        gaya_header(wp, len(kepala2), baris=4)
        for j, c in enumerate(prod["creator"][:K.JUMLAH_CREATOR_PER_PRODUK], 1):
            tulis_baris(wp, j + 4, [j] + [nilai_kolom(jd, c, k) for jd, k in K.PETA_KOLOM.items()])
        lebar_otomatis(wp)

    nama_keluar = os.path.join(
        K.DIR_HASIL,
        f"TOP50_CREATOR_{data['toko'].replace(' ', '_')}_{data['periode'].replace(' ', '_')}.xlsx")
    wb.save(nama_keluar)
    return nama_keluar, no, kode_tebakan


def main():
    berkas = sorted(glob.glob(os.path.join(K.DIR_MENTAH, "*.json")))
    if not berkas:
        print("Belum ada data mentah. Jalankan tarik_creator.py dulu.")
        sys.exit(1)
    for b in berkas:
        keluar, jumlah, tebakan = bangun(b)
        print(f"[OK] {os.path.basename(keluar)} - {jumlah} baris")
        if tebakan:
            print(f"\n[!] {len(tebakan)} kode 'Produk Top 10' masih TEBAKAN,")
            print("[!] diambil dari kata terakhir nama produk:")
            for pk, pid, kode in tebakan:
                print(f"      produk #{pk}: {kode}")
            print("[!]")
            print("[!] Betulkan di Excel: sheet RINGKASAN, kolom Kode (kuning).")
            print("[!] SIAP_PASTE ikut berubah sendiri. Simpan Excel-nya, lalu:")
            print("[!]     python simpan_kode.py")
            print("[!] supaya bulan depan tidak perlu dibetulkan lagi.")


if __name__ == "__main__":
    main()
